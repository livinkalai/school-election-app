import os
import sys
import json
import uuid
import webbrowser
import threading
import time
import socket
import ipaddress
import asyncio
import secrets  # For generating secure random key
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Tuple
from urllib.parse import urlparse
from fastapi import FastAPI, Request, Form, HTTPException, Response, Header
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from itsdangerous import URLSafeSerializer
import openpyxl
from openpyxl import Workbook
import requests # Needed for signaling existing process
import filelock # Needed for process locking
import tempfile # Needed for robust temporary file path
import hashlib # Needed for unique lock file name
from PIL import Image, ImageDraw
from pydantic import BaseModel
from license_manager import (
    LicenseError,
    initialize_runtime_license,
    enforce_license_or_raise,
    get_license_template_context,
    get_license_status,
)

# matplotlib / reportlab are imported inside export helpers only (avoids numpy init issues in PyInstaller onefile).

# Determine base paths for bundled vs. external files
if getattr(sys, 'frozen', False):
    # Running in a PyInstaller bundle
    APPLICATION_PATH = sys._MEIPASS
    EXTERNAL_FILES_PATH = os.path.dirname(sys.executable)
else:
    # Running in a normal Python environment
    APPLICATION_PATH = os.path.dirname(os.path.abspath(__file__))
    EXTERNAL_FILES_PATH = os.path.dirname(os.path.abspath(__file__))

# Initialize runtime licensing early (before config is applied).
try:
    initialize_runtime_license(
        application_path=APPLICATION_PATH,
        external_path=EXTERNAL_FILES_PATH,
        is_frozen=bool(getattr(sys, "frozen", False)),
    )
except LicenseError as license_boot_error:
    print(f"License validation failed: {license_boot_error}")
    sys.exit(1)

# Define the lock file path in a temporary directory for single instance enforcement
# Using a unique name derived from the app's base path for distinct instances
app_id = hashlib.sha256(EXTERNAL_FILES_PATH.encode()).hexdigest()[:10] # Short hash of path
LOCK_FILE_NAME = f"election_app_{app_id}.lock" # More unique name
LOCK_FILE_PATH = os.path.join(tempfile.gettempdir(), LOCK_FILE_NAME)

# Define settings / config paths (bundled in exe; optional flat override beside exe in dev)
SETTINGS_DIR = os.path.join(EXTERNAL_FILES_PATH, "settings")

def resolve_resource_file(filename: str, bundled_alternate_names: Optional[List[str]] = None) -> str:
    """Find a file: optional override beside exe, then bundled in exe, then dev settings folder."""
    search: List[str] = [
        os.path.join(EXTERNAL_FILES_PATH, filename),
        os.path.join(SETTINGS_DIR, filename),
    ]
    for alt in bundled_alternate_names or []:
        search.append(os.path.join(APPLICATION_PATH, "settings", alt))
    search.append(os.path.join(APPLICATION_PATH, "settings", filename))
    for path in search:
        if os.path.isfile(path):
            return path
    return os.path.join(SETTINGS_DIR, filename)

CONFIG_FILE = resolve_resource_file("config.json", ["config.example.json"])
CANDIDATES_FILE = resolve_resource_file("candidates.json")

# Dev mode: ensure settings directory exists for writable defaults
if not getattr(sys, "frozen", False) and not os.path.exists(SETTINGS_DIR):
    os.makedirs(SETTINGS_DIR)

def generate_secret_key():
    """Generate a secure random secret key."""
    return secrets.token_urlsafe(32)  # 32 bytes = 256 bits of entropy

def load_config():
    """Load configuration from bundled file, optional override beside exe, or dev settings/."""
    try:
        with open(CONFIG_FILE, 'r') as f:
            config = json.load(f)
            return config
    except FileNotFoundError:
        if getattr(sys, "frozen", False):
            raise Exception(
                "Configuration missing from application bundle. Rebuild the exe with settings/config.example.json."
            )
        default_config = {
            "school_name": "Your School Name",
            "logo_url": "/static/images/school_logo.svg",
            "background_url": "/static/images/school_bg.jpg",
            "admin_username": "admin",
            "admin_password": "admin123",
            "theme_name": "primary",
            "available_themes": [
                "primary",
                "secondary",
                "light",
                "warning",
                "info"
            ],
            "node_role": "secondary",
            "sync_secret": "",
            "machine_id": "",
            "lan_discovery": True,
            "lan_scan_cidrs": [],
        }
        with open(CONFIG_FILE, 'w') as f:
            json.dump(default_config, f, indent=4)
        return default_config
    except json.JSONDecodeError:
        raise Exception(f"Invalid config file: {CONFIG_FILE}")

def apply_sync_defaults(cfg: dict) -> None:
    """Ensure LAN sync keys exist (defaults: secondary, discovery on, empty secret)."""
    defaults = {
        "node_role": "secondary",
        "sync_secret": "",
        "machine_id": "",
        "lan_discovery": True,
        "lan_scan_cidrs": [],
    }
    for k, v in defaults.items():
        if k not in cfg:
            cfg[k] = v

def load_candidates() -> Dict[str, List[str]]:
    """Load candidates from bundled file, optional override beside exe, or dev settings/."""
    try:
        with open(CANDIDATES_FILE, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        if getattr(sys, "frozen", False):
            raise Exception(
                "Candidates missing from application bundle. Rebuild the exe with settings/candidates.json."
            )
        default_candidates = {
            "school_people_leader": [],
            "assistant_school_people_leader": []
        }
        with open(CANDIDATES_FILE, 'w') as f:
            json.dump(default_candidates, f, indent=4)
        return default_candidates
    except json.JSONDecodeError:
        raise Exception(f"Invalid candidates file: {CANDIDATES_FILE}")

# Load configuration
config = load_config()
apply_sync_defaults(config)

# In frozen deployments, school identity is license-controlled.
if getattr(sys, "frozen", False):
    status = get_license_status()
    licensed_school_name = str(status.get("school_name") or "").strip()
    if licensed_school_name:
        config["school_name"] = licensed_school_name

# Generate a secure secret key for session management
SECRET_KEY = generate_secret_key()

# Initialize serializer for session handling
serializer = URLSafeSerializer(SECRET_KEY)

def session_is_admin(data: Optional[dict]) -> bool:
    if not isinstance(data, dict):
        return False
    return data.get("is_admin") is True

def parse_session_data(cookie_val: Optional[str]) -> Optional[dict]:
    if not cookie_val:
        return None
    try:
        return serializer.loads(cookie_val)
    except Exception:
        return None

def session_template_vars(request: Request) -> dict:
    """Navbar: is_admin from cookie (password login). Machine roles are only primary | secondary."""
    data = parse_session_data(request.cookies.get("session"))
    license_ctx = get_license_template_context()
    branding = branding_static_urls()
    if not data:
        return {
            "is_admin": False,
            "is_primary_node": NODE_ROLE == "primary",
            **branding,
            **license_ctx,
        }
    return {
        "is_admin": data.get("is_admin") is True,
        "is_primary_node": NODE_ROLE == "primary",
        **branding,
        **license_ctx,
    }

# Get theme settings from config
THEME_NAME = config.get("theme_name", "primary")

# Get app host and port from config or defaults
APP_HOST = "0.0.0.0"  # Changed from 127.0.0.1 to 0.0.0.0 to allow network access
APP_PORT = 8001

app = FastAPI()
server_running = True
server = None  # Global server reference

# Static files: optional override beside exe, else bundled in exe, else dev project folder
EXTERNAL_STATIC_DIR = os.path.join(EXTERNAL_FILES_PATH, "static")
BUNDLED_STATIC_DIR = os.path.join(APPLICATION_PATH, "static")
if getattr(sys, "frozen", False):
    STATIC_DIR_PATH = EXTERNAL_STATIC_DIR if os.path.isdir(EXTERNAL_STATIC_DIR) else BUNDLED_STATIC_DIR
else:
    STATIC_DIR_PATH = EXTERNAL_STATIC_DIR
STATIC_IMAGES_PATH = os.path.join(STATIC_DIR_PATH, "images")
STATIC_CANDIDATES_PATH = os.path.join(STATIC_IMAGES_PATH, "candidates")

if not getattr(sys, "frozen", False):
    if not os.path.exists(STATIC_DIR_PATH):
        os.makedirs(STATIC_DIR_PATH)
    if not os.path.exists(STATIC_IMAGES_PATH):
        os.makedirs(STATIC_IMAGES_PATH)
    if not os.path.exists(STATIC_CANDIDATES_PATH):
        os.makedirs(STATIC_CANDIDATES_PATH)

def ensure_background_image():
    """Ensure the background image exists, create a default one if it doesn't."""
    background_url = config.get("background_url", "")
    if background_url:
        # Extract filename from URL
        background_file = os.path.basename(background_url)
        background_path = os.path.join(STATIC_IMAGES_PATH, background_file)
        
        # If background image doesn't exist, create a default one (dev / writable static only)
        if not os.path.exists(background_path) and not getattr(sys, "frozen", False):
            try:
                # Create a simple gradient background
                img = Image.new('RGB', (1920, 1080), color='#1a237e')
                draw = ImageDraw.Draw(img)
                
                # Draw some decorative elements
                for i in range(0, 1920, 50):
                    draw.line([(i, 0), (i, 1080)], fill='#283593', width=2)
                for i in range(0, 1080, 50):
                    draw.line([(0, i), (1920, i)], fill='#283593', width=2)
                
                # Save the image
                img.save(background_path)
                print(f"Created default background image at: {background_path}")
            except Exception as e:
                print(f"Error creating background image: {e}")

# Mount static files from bundle (exe) or project folder (dev)
app.mount("/static", StaticFiles(directory=STATIC_DIR_PATH), name="static")

# Ensure background image exists
ensure_background_image()

def static_url_with_version(url: str) -> str:
    """Append ?v=mtime so browser picks up replaced images beside the exe."""
    if not url or not url.startswith("/static/"):
        return url
    base_url, _, _ = url.partition("?")
    rel = base_url[len("/static/"):].lstrip("/").replace("/", os.sep)
    for base in (STATIC_DIR_PATH, BUNDLED_STATIC_DIR):
        path = os.path.join(base, rel)
        if os.path.isfile(path):
            return f"{base_url}?v={int(os.path.getmtime(path))}"
    return url


def branding_static_urls() -> Dict[str, str]:
    """Fresh cache-busted URLs for branding images (re-read mtime each request)."""
    return {
        "logo_url": static_url_with_version(config.get("logo_url", "")),
        "background_url": static_url_with_version(config.get("background_url", "")),
        "default_contact_image": static_url_with_version("/static/images/contact.png"),
    }

templates = Jinja2Templates(directory=os.path.join(APPLICATION_PATH, "templates"))

# File paths - external to exe
VOTES_FILE = os.path.join(EXTERNAL_FILES_PATH, "votes.xlsx")

# Admin credentials from config
ADMIN_USERNAME = config["admin_username"]
ADMIN_PASSWORD = config["admin_password"]
SCHOOL_NAME = config["school_name"]
LOGO_URL = static_url_with_version(config.get("logo_url", ""))
BACKGROUND_URL = static_url_with_version(config.get("background_url", ""))

MACHINE_ID = (config.get("machine_id") or "").strip() or socket.gethostname()
_raw_node_role = (config.get("node_role") or "secondary").strip().lower()
NODE_ROLE = "primary" if _raw_node_role == "primary" else "secondary"
SYNC_SECRET = (config.get("sync_secret") or "").strip()
LAN_DISCOVERY = bool(config.get("lan_discovery", True))
LAN_SCAN_CIDRS: List[str] = [str(c).strip() for c in (config.get("lan_scan_cidrs") or []) if str(c).strip()]
LAN_PEER_TIMEOUT = float(config.get("lan_peer_timeout", 0.35))
LAN_COLLECT_TIMEOUT = float(config.get("lan_collect_timeout", 4))
LAN_SCAN_WORKERS = int(config.get("lan_scan_workers", 64))
MAX_LAN_SCAN_HOSTS = int(config.get("max_lan_scan_hosts", 512))
# Bypass system proxy for LAN HTTP (avoids hangs on WiFi with no internet / captive portal).
_NO_PROXY = {"http": None, "https": None}

VOTES_HEADER_NEW = ["VoteId", "Timestamp", "Post", "CandidateName", "SourceMachine"]
PENDING_SYNC_FILE = os.path.join(EXTERNAL_FILES_PATH, "pending_sync.jsonl")
PEER_STATE_FILE = os.path.join(EXTERNAL_FILES_PATH, "peer_lan_state.json")
PEER_STATE_LOCK_FILE = os.path.join(EXTERNAL_FILES_PATH, "peer_lan_state.json.lock")
_last_sync_status: Dict[str, Any] = {"last_push_error": None, "last_collect": None}
_sync_stop = threading.Event()
_lan_primary_cache: Dict[str, Any] = {"urls": [], "ts": 0.0}
_lan_secondary_cache: Dict[str, Any] = {"urls": [], "ts": 0.0}
_LAN_PRIMARY_CACHE_TTL_SEC = 90.0
_LAN_SECONDARY_CACHE_TTL_SEC = 90.0


def _lan_request_kwargs(timeout: float) -> Dict[str, Any]:
    return {"timeout": timeout, "proxies": _NO_PROXY}


def get_export_branding_line() -> str:
    status = get_license_status()
    school = status.get("school_name") or SCHOOL_NAME
    developer = status.get("developer_name") or "EmpowerID"
    expires = status.get("expires_at") or "N/A"
    return f"Licensed to {school} | Valid until {expires} | Developed by {developer}"


def ensure_primary_node_or_forbidden() -> None:
    if NODE_ROLE != "primary":
        raise HTTPException(status_code=403, detail="Primary node required for this action.")


@app.middleware("http")
async def license_enforcement_middleware(request: Request, call_next):
    # Static assets are allowed so the expired page can render normally.
    if request.url.path.startswith("/static/"):
        return await call_next(request)
    try:
        enforce_license_or_raise()
    except LicenseError as e:
        if request.url.path.startswith("/api/"):
            return JSONResponse(status_code=403, content={"detail": str(e)})
        return templates.TemplateResponse(
            "license_expired.html",
            {
                "request": request,
                "school_name": SCHOOL_NAME,
                "theme_name": THEME_NAME,
                "error_message": str(e),
                **branding_static_urls(),
                **get_license_template_context(),
            },
            status_code=403,
        )
    return await call_next(request)

def get_local_ipv4_addresses() -> List[str]:
    """IPv4 addresses for this machine (excludes loopback). No DNS or internet required."""
    out: set = set()
    for probe_host in ("255.255.255.255", "10.255.255.255", "192.168.255.255"):
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
                s.settimeout(0.15)
                s.connect((probe_host, 1))
                ip = s.getsockname()[0]
                if ip and not str(ip).startswith("127."):
                    out.add(ip)
        except Exception:
            continue
    return list(out)

def get_scan_networks() -> List[ipaddress.IPv4Network]:
    """Networks to scan for peers. Uses lan_scan_cidrs if set; else /24 around each local IPv4."""
    if LAN_SCAN_CIDRS:
        nets: List[ipaddress.IPv4Network] = []
        for c in LAN_SCAN_CIDRS:
            try:
                nets.append(ipaddress.ip_network(c, strict=False))
            except Exception:
                continue
        return nets
    nets = []
    for ip in get_local_ipv4_addresses():
        try:
            nets.append(ipaddress.ip_interface(f"{ip}/24").network)
        except Exception:
            continue
    seen: set = set()
    out: List[ipaddress.IPv4Network] = []
    for n in nets:
        if n not in seen:
            seen.add(n)
            out.append(n)
    return out

def _probe_peer_base(base: str, want_role: str) -> Optional[str]:
    """Return base URL if peer responds with matching role and sync secret."""
    base = base.rstrip("/")
    try:
        r = requests.get(
            f"{base}/api/sync/peer",
            headers=_sync_headers(),
            **_lan_request_kwargs(LAN_PEER_TIMEOUT),
        )
        if r.status_code != 200:
            return None
        data = r.json()
        if data.get("app") != "school-election-app":
            return None
        role = str(data.get("node_role") or "").strip().lower()
        if role == want_role.strip().lower():
            return base
    except Exception:
        return None
    return None

def discover_bases_for_role(want_role: str) -> List[str]:
    """Scan LAN for other app instances with the given node_role (primary or secondary)."""
    wr = want_role.strip().lower()
    if wr not in ("primary", "secondary") or not SYNC_SECRET:
        return []
    networks = get_scan_networks()
    if not networks:
        return []
    self_ips = set(get_local_ipv4_addresses())
    candidates: List[str] = []
    for net in networks:
        try:
            n_hosts = int(net.num_addresses) - 2
        except Exception:
            n_hosts = 0
        if n_hosts > MAX_LAN_SCAN_HOSTS:
            print(
                f"LAN discovery: skipping {net} (>{MAX_LAN_SCAN_HOSTS} addresses). "
                "Set a smaller lan_scan_cidrs (e.g. 192.168.1.0/24)."
            )
            continue
        for host in net.hosts():
            if len(candidates) >= MAX_LAN_SCAN_HOSTS:
                break
            ip_str = str(host)
            if ip_str in self_ips:
                continue
            candidates.append(f"http://{ip_str}:{APP_PORT}")
        if len(candidates) >= MAX_LAN_SCAN_HOSTS:
            break
    found: List[str] = []
    if not candidates:
        return []
    with ThreadPoolExecutor(max_workers=max(4, min(LAN_SCAN_WORKERS, len(candidates)))) as ex:
        futures = {ex.submit(_probe_peer_base, b, wr): b for b in candidates}
        for fut in as_completed(futures):
            try:
                ok = fut.result()
                if ok:
                    found.append(ok)
            except Exception:
                continue
    return sorted(set(found))

def get_effective_primary_urls() -> List[str]:
    """LAN discovery only (no manual Primary URLs in config)."""
    if not LAN_DISCOVERY or not SYNC_SECRET:
        return []
    now = time.time()
    if (
        _lan_primary_cache["urls"]
        and (now - float(_lan_primary_cache["ts"])) < _LAN_PRIMARY_CACHE_TTL_SEC
    ):
        return list(_lan_primary_cache["urls"])
    urls = discover_bases_for_role("primary")
    _lan_primary_cache["urls"] = urls
    _lan_primary_cache["ts"] = now
    return list(urls)

def get_effective_secondary_urls(*, allow_lan_scan: bool = True) -> List[str]:
    """Known secondaries from peer state; optional cached LAN scan when none known."""
    if not LAN_DISCOVERY or not SYNC_SECRET:
        return []
    known = get_monitored_secondary_bases(scan_lan=False)
    if known:
        return known
    if not allow_lan_scan:
        return []
    now = time.time()
    if (
        _lan_secondary_cache["urls"]
        and (now - float(_lan_secondary_cache["ts"])) < _LAN_SECONDARY_CACHE_TTL_SEC
    ):
        return list(_lan_secondary_cache["urls"])
    urls = discover_bases_for_role("secondary")
    _lan_secondary_cache["urls"] = urls
    _lan_secondary_cache["ts"] = now
    return list(urls)

def _default_peer_state() -> dict:
    return {"secondaries": {}, "primaries": {}}

def load_peer_state() -> dict:
    lock = filelock.FileLock(PEER_STATE_LOCK_FILE)
    try:
        with lock.acquire(timeout=8):
            if not os.path.exists(PEER_STATE_FILE):
                return _default_peer_state()
            with open(PEER_STATE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, dict):
                return _default_peer_state()
            data.setdefault("secondaries", {})
            data.setdefault("primaries", {})
            if not isinstance(data["secondaries"], dict):
                data["secondaries"] = {}
            if not isinstance(data["primaries"], dict):
                data["primaries"] = {}
            return data
    except Exception:
        return _default_peer_state()

def save_peer_state(data: dict) -> None:
    lock = filelock.FileLock(PEER_STATE_LOCK_FILE)
    with lock.acquire(timeout=12):
        tmp = PEER_STATE_FILE + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        os.replace(tmp, PEER_STATE_FILE)

def merge_peer_section(section: str, base: str, updates: dict) -> None:
    base = base.rstrip("/")
    state = load_peer_state()
    m = state.setdefault(section, {})
    cur = dict(m.get(base, {}))
    cur.update(updates)
    m[base] = cur
    state[section] = m
    save_peer_state(state)

def extract_host_from_base(base: str) -> str:
    try:
        u = urlparse(base)
        return u.hostname or base
    except Exception:
        return base

def probe_peer_live(base: str) -> dict:
    try:
        r = requests.get(
            f"{base.rstrip('/')}/api/sync/peer",
            headers=_sync_headers(),
            **_lan_request_kwargs(0.8),
        )
        if r.status_code != 200:
            return {"ok": False, "error": f"HTTP {r.status_code}"}
        data = r.json()
        return {
            "ok": True,
            "machine_id": data.get("machine_id") or "",
            "node_role": str(data.get("node_role") or "").strip().lower(),
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}

def format_collect_summary(merged: dict) -> str:
    if not merged.get("last_collect_at"):
        return "Not collected yet"
    if merged.get("last_collect_error"):
        err = str(merged["last_collect_error"])
        return f"Failed: {err[:120]}{'…' if len(err) > 120 else ''}"
    n = merged.get("last_collect_merged")
    if n is None:
        return "—"
    return f"OK — {n} new vote(s) merged"

def format_push_summary(merged: dict) -> str:
    if not merged.get("last_push_at"):
        return "No push yet"
    if merged.get("last_push_ok") is True:
        return "OK"
    if merged.get("last_push_error"):
        err = str(merged["last_push_error"])
        return f"Failed: {err[:120]}{'…' if len(err) > 120 else ''}"
    return "Unknown"

def get_monitored_secondary_bases(scan_lan: bool) -> List[str]:
    bases = set()
    st = load_peer_state().get("secondaries", {})
    for k in st.keys():
        bases.add(k.rstrip("/"))
    if scan_lan and LAN_DISCOVERY and SYNC_SECRET:
        for b in discover_bases_for_role("secondary"):
            bases.add(b.rstrip("/"))
    return sorted(bases)

def get_monitored_primary_bases(scan_lan: bool) -> List[str]:
    bases = set()
    st = load_peer_state().get("primaries", {})
    for k in st.keys():
        bases.add(k.rstrip("/"))
    if scan_lan and LAN_DISCOVERY and SYNC_SECRET:
        for b in discover_bases_for_role("primary"):
            bases.add(b.rstrip("/"))
    return sorted(bases)

def build_secondary_monitor_rows(bases: List[str]) -> List[dict]:
    if not bases:
        return []
    state = load_peer_state()
    secondaries = dict(state.setdefault("secondaries", {}))
    prior = {k: dict(v) for k, v in secondaries.items()}
    rows_out: List[dict] = []

    def work(base: str) -> Tuple[str, dict, dict]:
        base = base.rstrip("/")
        live = probe_peer_live(base)
        pr = dict(prior.get(base, {}))
        upd = {
            "last_probe_at": datetime.now().isoformat(),
            "last_probe_ok": live["ok"],
            "last_probe_error": None if live["ok"] else live.get("error", ""),
        }
        if live.get("machine_id"):
            upd["machine_id"] = live["machine_id"]
        if live.get("node_role"):
            upd["node_role"] = live["node_role"]
        merged = {**pr, **upd}
        display = {
            "ip": extract_host_from_base(base),
            "base": base,
            "machine_id": merged.get("machine_id") or "—",
            "app_online": live["ok"],
            "node_role": merged.get("node_role") or "—",
            "last_probe_at": merged.get("last_probe_at"),
            "last_collect_at": merged.get("last_collect_at"),
            "last_collect_merged": merged.get("last_collect_merged"),
            "last_collect_error": merged.get("last_collect_error"),
            "collect_summary": format_collect_summary(merged),
        }
        return base, merged, display

    with ThreadPoolExecutor(max_workers=max(4, min(32, len(bases)))) as ex:
        futs = [ex.submit(work, b) for b in bases]
        for fut in as_completed(futs):
            base, merged, display = fut.result()
            secondaries[base] = merged
            rows_out.append(display)
    state["secondaries"] = secondaries
    save_peer_state(state)
    rows_out.sort(key=lambda r: r["ip"])
    return rows_out

def build_primary_monitor_rows(bases: List[str]) -> List[dict]:
    if not bases:
        return []
    state = load_peer_state()
    primaries = dict(state.setdefault("primaries", {}))
    prior = {k: dict(v) for k, v in primaries.items()}
    rows_out: List[dict] = []

    def work(base: str) -> Tuple[str, dict, dict]:
        base = base.rstrip("/")
        live = probe_peer_live(base)
        pr = dict(prior.get(base, {}))
        upd = {
            "last_probe_at": datetime.now().isoformat(),
            "last_probe_ok": live["ok"],
            "last_probe_error": None if live["ok"] else live.get("error", ""),
        }
        if live.get("machine_id"):
            upd["machine_id"] = live["machine_id"]
        if live.get("node_role"):
            upd["node_role"] = live["node_role"]
        merged = {**pr, **upd}
        display = {
            "ip": extract_host_from_base(base),
            "base": base,
            "machine_id": merged.get("machine_id") or "—",
            "app_online": live["ok"],
            "node_role": merged.get("node_role") or "—",
            "last_probe_at": merged.get("last_probe_at"),
            "last_push_at": merged.get("last_push_at"),
            "last_push_ok": merged.get("last_push_ok"),
            "last_push_error": merged.get("last_push_error"),
            "push_summary": format_push_summary(merged),
        }
        return base, merged, display

    with ThreadPoolExecutor(max_workers=max(4, min(32, len(bases)))) as ex:
        futs = [ex.submit(work, b) for b in bases]
        for fut in as_completed(futs):
            base, merged, display = fut.result()
            primaries[base] = merged
            rows_out.append(display)
    state["primaries"] = primaries
    save_peer_state(state)
    rows_out.sort(key=lambda r: r["ip"])
    return rows_out

def _normalize_base_url(url: str) -> str:
    u = (url or "").strip().rstrip("/")
    return u

def migrate_votes_file_if_needed() -> None:
    """Migrate legacy 3-column votes.xlsx to VoteId + SourceMachine schema."""
    if not os.path.exists(VOTES_FILE):
        return
    lock_file = f"{VOTES_FILE}.lock"
    lock = filelock.FileLock(lock_file)
    try:
        with lock.acquire(timeout=10):
            wb = openpyxl.load_workbook(VOTES_FILE)
            ws = wb.active
            headers = [c.value for c in ws[1]]
            if not headers:
                wb.close()
                return
            first = str(headers[0] or "").strip()
            if first == "VoteId" and len(headers) >= 5:
                wb.close()
                return
            if first == "Timestamp" and len(headers) >= 3:
                rows_out: List[List[Any]] = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    ts, post, cand = row[0], row[1], row[2]
                    rows_out.append([
                        str(uuid.uuid4()),
                        ts,
                        post,
                        cand,
                        MACHINE_ID or "legacy",
                    ])
                wb.close()
                out_wb = Workbook()
                out_ws = out_wb.active
                out_ws.append(VOTES_HEADER_NEW)
                for r in rows_out:
                    out_ws.append(r)
                for cell in out_ws[1]:
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                for col_letter, w in zip(["A", "B", "C", "D", "E"], [36, 20, 30, 30, 24]):
                    out_ws.column_dimensions[col_letter].width = w
                out_wb.save(VOTES_FILE)
                out_wb.close()
                print("Migrated votes.xlsx to multi-LAN schema (VoteId, SourceMachine).")
                return
            wb.close()
    except filelock.Timeout:
        raise Exception("Could not migrate votes file (lock timeout).")
    finally:
        if os.path.exists(lock_file):
            try:
                os.remove(lock_file)
            except Exception:
                pass

def initialize_votes_file():
    """Create votes.xlsx if it doesn't exist."""
    try:
        if not os.path.exists(VOTES_FILE):
            print(f"Creating votes file at: {VOTES_FILE}")
            wb = Workbook()
            ws = wb.active
            ws.append(VOTES_HEADER_NEW)
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            for col_letter, w in zip(["A", "B", "C", "D", "E"], [36, 20, 30, 30, 24]):
                ws.column_dimensions[col_letter].width = w
            wb.save(VOTES_FILE)
            print("Votes file created successfully!")
        else:
            print(f"Votes file already exists at: {VOTES_FILE}")
            migrate_votes_file_if_needed()
    except PermissionError:
        raise Exception("Cannot create votes file. Please make sure it's not open in another program.")
    except Exception as e:
        raise Exception(f"Error creating votes file: {str(e)}")

def create_batch_files():
    """Create stop.bat and refresh.bat files if they don't exist."""
    try:
        # Create stop.bat
        stop_bat_path = os.path.join(EXTERNAL_FILES_PATH, "stop.bat")
        if not os.path.exists(stop_bat_path):
            print("Creating stop.bat...")
            with open(stop_bat_path, 'w') as f:
                f.write('@echo off\n')
                f.write('echo Stopping School Election App...\n')
                f.write('taskkill /F /IM school-election-app.exe /T\n')
                f.write('if %ERRORLEVEL% EQU 0 (\n')
                f.write('    echo Application stopped successfully.\n')
                f.write(') else (\n')
                f.write('    echo No running instance found or application already stopped.\n')
                f.write(')\n')
                f.write('timeout /t 2 >nul\n')
            print("stop.bat created successfully!")

        # Create refresh.bat
        refresh_bat_path = os.path.join(EXTERNAL_FILES_PATH, "refresh.bat")
        if not os.path.exists(refresh_bat_path):
            print("Creating refresh.bat...")
            with open(refresh_bat_path, 'w') as f:
                f.write('@echo off\n')
                f.write('echo Stopping School Election App...\n')
                f.write('taskkill /F /IM school-election-app.exe /T\n')
                f.write('if %ERRORLEVEL% EQU 0 (\n')
                f.write('    echo Application stopped successfully.\n')
                f.write(') else (\n')
                f.write('    echo No running instance found or application already stopped.\n')
                f.write(')\n')
                f.write('timeout /t 2 >nul\n')
                f.write('\n')
                f.write('echo Starting School Election App...\n')
                f.write('start "" "school-election-app.exe"\n')
                f.write('echo Application started successfully.\n')
            print("refresh.bat created successfully!")
    except Exception as e:
        print(f"Warning: Could not create batch files: {str(e)}")

def append_vote_row(vote_id: str, timestamp: datetime, post: str, candidate: str, source_machine: str) -> None:
    """Append one vote row (must hold votes file lock externally)."""
    wb = openpyxl.load_workbook(VOTES_FILE)
    ws = wb.active
    ws.append([vote_id, timestamp, post, candidate, source_machine])
    wb.save(VOTES_FILE)

def save_vote(post: str, candidate: str) -> Tuple[str, datetime]:
    """Save vote to Excel file. Returns (vote_id, timestamp)."""
    if not post or not candidate:
        raise Exception("Invalid vote: post and candidate cannot be empty")

    vote_id = str(uuid.uuid4())
    timestamp = datetime.now()
    source_machine = MACHINE_ID
    max_retries = 3
    retry_delay = 1
    lock_file = f"{VOTES_FILE}.lock"
    lock = filelock.FileLock(lock_file)

    try:
        with lock.acquire(timeout=5):
            for attempt in range(max_retries):
                try:
                    if not isinstance(timestamp, datetime) or not isinstance(post, str) or not isinstance(candidate, str):
                        raise Exception("Invalid vote data format")
                    append_vote_row(vote_id, timestamp, post, candidate, source_machine)
                    return vote_id, timestamp
                except PermissionError:
                    if attempt < max_retries - 1:
                        time.sleep(retry_delay)
                        continue
                    raise Exception("Cannot save vote. Please make sure the votes file is not open in another program.")
                except Exception as e:
                    if attempt < max_retries - 1:
                        time.sleep(retry_delay)
                        continue
                    raise Exception(f"Error saving vote: {str(e)}")
    except filelock.Timeout:
        raise Exception("Could not acquire lock on votes file. Please try again.")
    finally:
        if os.path.exists(lock_file):
            try:
                os.remove(lock_file)
            except Exception:
                pass

def get_voted_posts(session_data: dict) -> List[str]:
    """Get list of posts already voted in current session."""
    return session_data.get("voted_posts", [])

def get_results() -> Dict[str, Dict[str, int]]:
    """Calculate voting results by post and candidate only."""
    # Initialize results with all candidates having 0 votes
    results = {}
    candidates = load_candidates()
    
    # Initialize all posts and candidates with 0 votes
    for post, candidate_list in candidates.items():
        results[post] = {}
        for candidate in candidate_list:
            results[post][candidate] = 0
    
    # If no votes file exists, return initialized results
    if not os.path.exists(VOTES_FILE):
        return results
    
    max_retries = 3
    retry_delay = 1  # seconds
    
    for attempt in range(max_retries):
        try:
            wb = openpyxl.load_workbook(VOTES_FILE, read_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(h or "").strip() for h in header_row]
            try:
                pi = int(headers.index("Post"))
                ci = int(headers.index("CandidateName"))
            except ValueError:
                pi, ci = 2, 3
                if len(headers) >= 3 and str(headers[0] or "").strip() == "Timestamp":
                    pi, ci = 1, 2
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= max(pi, ci):
                    continue
                post = row[pi]
                candidate = row[ci]
                if post is None or candidate is None:
                    continue
                post = str(post).strip()
                candidate = str(candidate).strip()
                if not post or not candidate:
                    continue
                if post not in results:
                    results[post] = {}
                if candidate not in results[post]:
                    results[post][candidate] = 0
                results[post][candidate] += 1
            wb.close()
            return results
        except PermissionError:
            if attempt < max_retries - 1:  # If not the last attempt
                time.sleep(retry_delay)
                continue
            raise Exception("Cannot read results. Please make sure the votes file is not open in another program.")
        except Exception as e:
            if attempt < max_retries - 1:  # If not the last attempt
                time.sleep(retry_delay)
                continue
            raise Exception(f"Error reading results: {str(e)}")

def cleanup_votes_file():
    """Remove blank rows from votes.xlsx (safe rewrite — avoids openpyxl delete_rows issues)."""
    if not os.path.exists(VOTES_FILE):
        return

    lock_file = f"{VOTES_FILE}.lock"
    lock = filelock.FileLock(lock_file)

    try:
        with lock.acquire(timeout=5):
            wb = openpyxl.load_workbook(VOTES_FILE, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if not rows:
                return

            header = rows[0]
            header_label = str(header[0] or "").strip() if header else ""
            valid_rows: List[tuple] = [header]

            for row in rows[1:]:
                if not row:
                    continue
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue
                valid_rows.append(row)

            if len(valid_rows) == len(rows):
                return

            out_wb = Workbook()
            out_ws = out_wb.active
            for row in valid_rows:
                out_ws.append(list(row))
            out_wb.save(VOTES_FILE)
            out_wb.close()
    except filelock.Timeout:
        raise Exception("Could not acquire lock on votes file for cleanup.")
    finally:
        if os.path.exists(lock_file):
            try:
                os.remove(lock_file)
            except Exception:
                pass

def sort_results_for_display(results: Dict[str, Dict[str, int]]) -> Dict[str, List[Tuple[str, int]]]:
    """Sort candidates by vote count (high to low) for the results template."""
    return {
        post: sorted(cand_map.items(), key=lambda item: item[1], reverse=True)
        for post, cand_map in results.items()
    }

def get_candidate_image(candidate_name: str) -> str:
    """Get the path to a candidate's image, or return default if not found."""
    candidate_image = os.path.join("candidates", f"{candidate_name}.png")
    candidate_image_path = os.path.join(STATIC_CANDIDATES_PATH, f"{candidate_name}.png")

    if os.path.exists(candidate_image_path):
        return static_url_with_version(f"/static/images/{candidate_image}")

    return static_url_with_version("/static/images/contact.png")

def _sync_headers() -> Dict[str, str]:
    h: Dict[str, str] = {}
    if SYNC_SECRET:
        h["X-Sync-Secret"] = SYNC_SECRET
    return h

def append_pending_sync(payload: dict) -> None:
    line = json.dumps(payload, ensure_ascii=False) + "\n"
    with open(PENDING_SYNC_FILE, "a", encoding="utf-8") as f:
        f.write(line)

def push_vote_to_primaries(vote_id: str, ts: datetime, post: str, candidate: str, source_machine: str) -> None:
    targets = get_effective_primary_urls()
    if NODE_ROLE != "secondary" or not SYNC_SECRET or not targets:
        return
    body = {
        "vote_id": vote_id,
        "timestamp": ts.isoformat(),
        "post": post,
        "candidate": candidate,
        "source_machine": source_machine,
    }
    headers = _sync_headers()
    errs = []
    ok_count = 0
    now_iso = datetime.now().isoformat()
    for base in targets:
        url = f"{base}/api/votes/ingest"
        try:
            r = requests.post(
                url, json=body, headers=headers, **_lan_request_kwargs(8),
            )
            if r.status_code in (200, 201):
                ok_count += 1
                merge_peer_section("primaries", base, {
                    "last_push_at": now_iso,
                    "last_push_ok": True,
                    "last_push_error": None,
                })
            else:
                merge_peer_section("primaries", base, {
                    "last_push_at": now_iso,
                    "last_push_ok": False,
                    "last_push_error": f"HTTP {r.status_code}",
                })
                errs.append(f"{base}: HTTP {r.status_code}")
        except Exception as e:
            merge_peer_section("primaries", base, {
                "last_push_at": now_iso,
                "last_push_ok": False,
                "last_push_error": str(e),
            })
            errs.append(f"{base}: {e}")
    if ok_count == 0:
        _last_sync_status["last_push_error"] = "; ".join(errs) if errs else "no primaries reachable"
        append_pending_sync(body)
    else:
        _last_sync_status["last_push_error"] = (
            None if ok_count == len(targets) else "Partial: " + "; ".join(errs)
        )

def process_pending_sync_queue_once() -> None:
    targets = get_effective_primary_urls()
    if NODE_ROLE != "secondary" or not SYNC_SECRET or not targets:
        return
    if not os.path.exists(PENDING_SYNC_FILE):
        return
    try:
        with open(PENDING_SYNC_FILE, "r", encoding="utf-8") as f:
            lines = f.readlines()
    except Exception:
        return
    if not lines:
        return
    remaining: List[str] = []
    headers = _sync_headers()
    for line in lines:
        line = line.strip()
        if not line:
            continue
        try:
            body = json.loads(line)
        except json.JSONDecodeError:
            remaining.append(line + "\n")
            continue
        ok_any = False
        now_iso = datetime.now().isoformat()
        for base in targets:
            url = f"{base}/api/votes/ingest"
            try:
                r = requests.post(
                url, json=body, headers=headers, **_lan_request_kwargs(8),
            )
                if r.status_code in (200, 201):
                    ok_any = True
                    merge_peer_section("primaries", base, {
                        "last_push_at": now_iso,
                        "last_push_ok": True,
                        "last_push_error": None,
                    })
                    break
            except Exception:
                pass
        if not ok_any:
            remaining.append(json.dumps(body, ensure_ascii=False) + "\n")
    with open(PENDING_SYNC_FILE, "w", encoding="utf-8") as f:
        f.writelines(remaining)
    if not remaining:
        _last_sync_status["last_push_error"] = None

def sync_retry_worker() -> None:
    while not _sync_stop.is_set():
        time.sleep(30)
        try:
            process_pending_sync_queue_once()
        except Exception as e:
            print(f"Pending sync retry error: {e}")


def license_watchdog_worker() -> None:
    """Periodic runtime license verification; shuts down server when expired."""
    while not _sync_stop.is_set():
        time.sleep(300)
        try:
            enforce_license_or_raise()
        except LicenseError as e:
            print(f"License watchdog stop: {e}")
            global server
            if server:
                server.should_exit = True
            break

def parse_iso_timestamp(s: Optional[str]) -> datetime:
    if not s:
        return datetime.now()
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except Exception:
        return datetime.now()

class VoteIngestBody(BaseModel):
    vote_id: str
    timestamp: Optional[str] = None
    post: str
    candidate: str
    source_machine: str = ""

def merge_remote_votes(rows: List[dict]) -> Tuple[int, int]:
    """Insert remote vote dicts if vote_id not present. Returns (inserted, skipped)."""
    if not rows:
        return 0, 0
    inserted = 0
    skipped = 0
    lock_file = f"{VOTES_FILE}.lock"
    lock = filelock.FileLock(lock_file)
    try:
        with lock.acquire(timeout=20):
            wb = openpyxl.load_workbook(VOTES_FILE)
            ws = wb.active
            existing: set = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    existing.add(str(row[0]).strip())
            for item in rows:
                vid = str(item.get("vote_id") or "").strip()
                if not vid:
                    skipped += 1
                    continue
                if vid in existing:
                    skipped += 1
                    continue
                ts_raw = item.get("timestamp")
                if isinstance(ts_raw, datetime):
                    ts = ts_raw
                else:
                    ts = parse_iso_timestamp(str(ts_raw) if ts_raw is not None else "")
                post = str(item.get("post") or "")
                cand = str(item.get("candidate") or "")
                src = str(item.get("source_machine") or "")
                if not post or not cand:
                    skipped += 1
                    continue
                ws.append([vid, ts, post, cand, src])
                existing.add(vid)
                inserted += 1
            wb.save(VOTES_FILE)
    except filelock.Timeout:
        raise RuntimeError("votes file busy")
    finally:
        if os.path.exists(lock_file):
            try:
                os.remove(lock_file)
            except Exception:
                pass
    return inserted, skipped

def verify_sync_secret_header(x_sync_secret: Optional[str]) -> bool:
    if not SYNC_SECRET:
        return False
    return (x_sync_secret or "").strip() == SYNC_SECRET

@app.get("/api/sync/peer")
async def api_sync_peer(x_sync_secret: Optional[str] = Header(None)):
    """LAN discovery: identifies this node when X-Sync-Secret matches."""
    if not SYNC_SECRET:
        raise HTTPException(status_code=503, detail="sync_secret not configured on this node")
    if not verify_sync_secret_header(x_sync_secret):
        raise HTTPException(status_code=403, detail="invalid sync secret")
    return {
        "app": "school-election-app",
        "machine_id": MACHINE_ID,
        "node_role": NODE_ROLE,
    }

def record_secondary_from_request(request: Request, source_machine: str = "") -> None:
    """Remember Secondary base URL when it pushes a vote (for Collect Now without full LAN scan)."""
    if not request.client or not request.client.host:
        return
    host = request.client.host.strip()
    if not host or host in ("127.0.0.1", "::1"):
        return
    base = f"http://{host}:{APP_PORT}".rstrip("/")
    updates: Dict[str, Any] = {"last_seen_push_at": datetime.now().isoformat()}
    if source_machine:
        updates["machine_id"] = source_machine
    merge_peer_section("secondaries", base, updates)

@app.post("/api/votes/ingest")
async def api_votes_ingest(body: VoteIngestBody, request: Request, x_sync_secret: Optional[str] = Header(None)):
    if not SYNC_SECRET:
        raise HTTPException(status_code=503, detail="sync_secret not configured on this node")
    if NODE_ROLE != "primary":
        raise HTTPException(status_code=403, detail="ingest only accepted on primary nodes")
    if not verify_sync_secret_header(x_sync_secret):
        raise HTTPException(status_code=403, detail="invalid sync secret")
    if not str(body.post or "").strip() or not str(body.candidate or "").strip():
        raise HTTPException(status_code=400, detail="post and candidate required")
    payload = {
        "vote_id": body.vote_id,
        "timestamp": body.timestamp,
        "post": body.post,
        "candidate": body.candidate,
        "source_machine": body.source_machine or "remote",
    }
    try:
        ins, _skipped = merge_remote_votes([payload])
    except RuntimeError:
        raise HTTPException(status_code=503, detail="votes file busy")
    record_secondary_from_request(request, str(body.source_machine or ""))
    if ins == 0:
        return {"ok": True, "duplicate": True}
    return {"ok": True, "duplicate": False}

@app.get("/api/votes/export")
async def api_votes_export(x_sync_secret: Optional[str] = Header(None)):
    if not SYNC_SECRET:
        raise HTTPException(status_code=503, detail="sync_secret not configured on this node")
    if not verify_sync_secret_header(x_sync_secret):
        raise HTTPException(status_code=403, detail="invalid sync secret")
    rows: List[dict] = []
    if not os.path.exists(VOTES_FILE):
        return {"votes": rows}
    wb = openpyxl.load_workbook(VOTES_FILE, read_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h or "").strip() for h in header_row]
    try:
        vi = headers.index("VoteId")
        ti = headers.index("Timestamp")
        pi = headers.index("Post")
        ci = headers.index("CandidateName")
        si = headers.index("SourceMachine")
    except ValueError:
        wb.close()
        return {"votes": rows}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= max(vi, ti, pi, ci, si):
            continue
        ts = row[ti]
        if hasattr(ts, "isoformat"):
            ts_str = ts.isoformat()
        else:
            ts_str = str(ts)
        rows.append({
            "vote_id": str(row[vi] or ""),
            "timestamp": ts_str,
            "post": str(row[pi] or ""),
            "candidate": str(row[ci] or ""),
            "source_machine": str(row[si] or ""),
        })
    wb.close()
    return {"votes": rows}


def _collect_votes_from_one_secondary(base: str) -> Tuple[str, str, int, int]:
    """Fetch and merge votes from one Secondary. Returns (base, summary_line, inserted, skipped)."""
    base = base.rstrip("/")
    url = f"{base}/api/votes/export"
    now_iso = datetime.now().isoformat()
    headers = _sync_headers()
    try:
        r = requests.get(url, headers=headers, **_lan_request_kwargs(LAN_COLLECT_TIMEOUT))
        if r.status_code != 200:
            merge_peer_section("secondaries", base, {
                "last_collect_at": now_iso,
                "last_collect_error": f"HTTP {r.status_code}",
            })
            return base, f"{base}: HTTP {r.status_code}", 0, 0
        payload = r.json()
        votes = payload.get("votes") or []
        try:
            ins, sk = merge_remote_votes(votes)
        except RuntimeError:
            merge_peer_section("secondaries", base, {
                "last_collect_at": now_iso,
                "last_collect_error": "votes file busy",
            })
            return base, f"{base}: votes file busy", 0, 0
        merge_peer_section("secondaries", base, {
            "last_collect_at": now_iso,
            "last_collect_merged": ins,
            "last_collect_skipped": sk,
            "last_collect_error": None,
        })
        return base, f"{base}: merged {ins} new, {sk} skipped/duplicate", ins, sk
    except Exception as e:
        merge_peer_section("secondaries", base, {
            "last_collect_at": now_iso,
            "last_collect_error": str(e),
        })
        return base, f"{base}: {e}", 0, 0


def run_collect_from_secondaries(secondaries: List[str]) -> Dict[str, Any]:
    """Collect votes from all known Secondaries in parallel (blocking — run off event loop)."""
    summary: List[str] = []
    total_in = 0
    total_skip = 0
    if not secondaries:
        return {"summary": summary, "total_in": total_in, "total_skip": total_skip}
    workers = max(1, min(16, len(secondaries)))
    with ThreadPoolExecutor(max_workers=workers) as ex:
        for base, line, ins, sk in ex.map(_collect_votes_from_one_secondary, secondaries):
            summary.append(line)
            total_in += ins
            total_skip += sk
    return {"summary": summary, "total_in": total_in, "total_skip": total_skip}


@app.post("/admin/collect-now")
async def admin_collect_now(request: Request):
    session = request.cookies.get("session")
    if not session:
        return RedirectResponse(url="/", status_code=303)
    try:
        data = serializer.loads(session)
        if not session_is_admin(data):
            return RedirectResponse(url="/", status_code=303)
    except Exception:
        return RedirectResponse(url="/", status_code=303)
    if NODE_ROLE != "primary":
        raise HTTPException(status_code=403, detail="Primary node required.")
    if not SYNC_SECRET:
        return RedirectResponse(url="/results?msg=no_secret", status_code=303)
    # Prefer known peers (no subnet scan). If none yet, use cached LAN discovery once.
    secondaries = get_effective_secondary_urls(allow_lan_scan=False)
    if not secondaries:
        secondaries = get_effective_secondary_urls(allow_lan_scan=True)
    if not secondaries:
        return RedirectResponse(url="/results?msg=no_secondaries", status_code=303)
    collect_result = await asyncio.to_thread(run_collect_from_secondaries, secondaries)
    _last_sync_status["last_collect"] = collect_result
    return RedirectResponse(url="/results?msg=collect_done", status_code=303)

@app.on_event("startup")
async def startup_pending_sync():
    t = threading.Thread(target=sync_retry_worker, daemon=True)
    t.start()
    lt = threading.Thread(target=license_watchdog_worker, daemon=True)
    lt.start()

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    """Show home page with Admin/Student buttons."""
    return templates.TemplateResponse("index.html", {
        "request": request,
        "school_name": SCHOOL_NAME,
        "logo_url": LOGO_URL,
        "theme_name": THEME_NAME,
        "background_url": BACKGROUND_URL,
        **session_template_vars(request),
    })

@app.get("/admin-login", response_class=HTMLResponse)
async def admin_login_page(request: Request):
    """Show admin login page."""
    return templates.TemplateResponse("admin_login.html", {
        "request": request,
        "school_name": SCHOOL_NAME,
        "logo_url": LOGO_URL,
        "theme_name": THEME_NAME,
        "background_url": BACKGROUND_URL,
        **session_template_vars(request),
    })

@app.post("/admin-login")
async def admin_login(request: Request, password: str = Form(...)):
    """Handle admin login."""
    if password == ADMIN_PASSWORD:
        session_data = serializer.dumps({"is_admin": True})
        response = RedirectResponse(url="/results", status_code=303)
        response.set_cookie(key="session", value=session_data)
        return response
    return RedirectResponse(url="/admin-login?error=1", status_code=303)

@app.get("/student-voting", response_class=HTMLResponse)
async def student_voting(request: Request):
    """Start student voting session."""
    session_data = serializer.dumps({"is_student": True, "voted_posts": []})
    response = RedirectResponse(url="/posts", status_code=303)
    response.set_cookie(key="session", value=session_data)
    return response

@app.get("/posts", response_class=HTMLResponse)
async def show_posts(request: Request):
    """Display available posts."""
    session = request.cookies.get("session")
    if not session:
        return RedirectResponse(url="/", status_code=303)
    
    try:
        data = serializer.loads(session)
        is_student = data.get("is_student", False)
        
        if session_is_admin(data):
            return RedirectResponse(url="/results", status_code=303)
        
        if not is_student:
            return RedirectResponse(url="/", status_code=303)
            
    except:
        return RedirectResponse(url="/", status_code=303)
    
    candidates = load_candidates()
    voted_posts = get_voted_posts(data)
    
    return templates.TemplateResponse(
        "posts.html",
        {
            "request": request,
            "posts": candidates,
            "voted_posts": voted_posts,
            "is_student": True,
            "school_name": SCHOOL_NAME,
            "logo_url": LOGO_URL,
            "theme_name": THEME_NAME,
            "background_url": BACKGROUND_URL,
            **session_template_vars(request),
        }
    )

@app.get("/vote/{post}", response_class=HTMLResponse)
async def vote_page(request: Request, post: str):
    """Show voting page for selected post."""
    session = request.cookies.get("session")
    if not session:
        return RedirectResponse(url="/", status_code=303)
    
    try:
        data = serializer.loads(session)
        is_student = data.get("is_student", False)
        if not is_student:
            return RedirectResponse(url="/", status_code=303)
    except:
        return RedirectResponse(url="/", status_code=303)
    
    # Check if already voted for this post
    if post in get_voted_posts(data):
        return RedirectResponse(url="/posts", status_code=303)
    
    candidates = load_candidates()
    if post not in candidates:
        raise HTTPException(status_code=404, detail="Post not found")
    
    # Get image paths for each candidate
    candidate_images = {candidate: get_candidate_image(candidate) for candidate in candidates[post]}
    
    return templates.TemplateResponse(
        "vote.html",
        {
            "request": request,
            "post": post,
            "candidates": candidates[post],
            "candidate_images": candidate_images,
            "is_student": True,
            "school_name": SCHOOL_NAME,
            "logo_url": LOGO_URL,
            "theme_name": THEME_NAME,
            "background_url": BACKGROUND_URL,
            **session_template_vars(request),
        }
    )

@app.post("/vote/{post}")
async def submit_vote(request: Request, post: str, candidate: str = Form(...)):
    """Process vote submission."""
    session = request.cookies.get("session")
    if not session:
        return RedirectResponse(url="/", status_code=303)
    
    try:
        data = serializer.loads(session)
        is_student = data.get("is_student", False)
        if not is_student:
            return RedirectResponse(url="/", status_code=303)
    except:
        return RedirectResponse(url="/", status_code=303)
    
    # Check if already voted for this post
    if post in get_voted_posts(data):
        return RedirectResponse(url="/posts", status_code=303)
    
    try:
        vote_id, ts = save_vote(post, candidate)

        def _push_async():
            try:
                push_vote_to_primaries(vote_id, ts, post, candidate, MACHINE_ID)
            except Exception as ex:
                print(f"LAN sync push error: {ex}")

        threading.Thread(target=_push_async, daemon=True).start()

        # Update session with voted post
        voted_posts = get_voted_posts(data)
        voted_posts.append(post)
        session_data = serializer.dumps({"is_student": True, "voted_posts": voted_posts})
        response = RedirectResponse(url="/posts", status_code=303)
        response.set_cookie(key="session", value=session_data)
        
        return response
    except Exception as e:
        # If there's an error saving the vote, show error page
        return templates.TemplateResponse(
            "error.html",
            {
                "request": request,
                "error_message": str(e),
                "school_name": SCHOOL_NAME,
                "logo_url": LOGO_URL,
                "theme_name": THEME_NAME,
                "background_url": BACKGROUND_URL,
                **session_template_vars(request),
            }
        )

@app.get("/reset-voting")
async def reset_voting(request: Request):
    """Reset voting session for next voter while staying in student mode."""
    session = request.cookies.get("session")
    if not session:
        return RedirectResponse(url="/", status_code=303)
    
    try:
        data = serializer.loads(session)
        is_student = data.get("is_student", False)
        if not is_student:
            return RedirectResponse(url="/", status_code=303)
    except:
        return RedirectResponse(url="/", status_code=303)
    
    # Create new session with empty voted posts
    session_data = serializer.dumps({"is_student": True, "voted_posts": []})
    response = RedirectResponse(url="/posts", status_code=303)
    response.set_cookie(key="session", value=session_data)
    return response

@app.get("/end-voting")
async def end_voting():
    """End student voting session and return to home page."""
    response = RedirectResponse(url="/")
    response.delete_cookie("session")
    return response

@app.get("/results", response_class=HTMLResponse)
async def results(request: Request):
    """Show voting results."""
    session = request.cookies.get("session")
    if not session:
        return RedirectResponse(url="/", status_code=303)
    
    try:
        data = serializer.loads(session)
        if not session_is_admin(data):
            return RedirectResponse(url="/", status_code=303)
            
        # Clean up any empty rows before showing results
        cleanup_votes_file()
        
        vote_results = get_results()
        sorted_results = sort_results_for_display(vote_results)
        msg = request.query_params.get("msg")
        is_primary = NODE_ROLE == "primary"
        return templates.TemplateResponse("results.html", {
            "request": request,
            "results": vote_results,
            "sorted_results": sorted_results,
            "school_name": SCHOOL_NAME,
            "logo_url": LOGO_URL,
            "theme_name": THEME_NAME,
            "background_url": BACKGROUND_URL,
            "is_primary": is_primary,
            "sync_secret_set": bool(SYNC_SECRET),
            "last_push_error": _last_sync_status.get("last_push_error") if is_primary else None,
            "last_collect": _last_sync_status.get("last_collect"),
            "results_msg": msg,
            **session_template_vars(request),
        })
    except Exception as e:
        print(f"Results page error: {e}")
        return templates.TemplateResponse("error.html", {
            "request": request,
            "error_message": f"Could not load results: {e}",
            "school_name": SCHOOL_NAME,
            "logo_url": LOGO_URL,
            "theme_name": THEME_NAME,
            "background_url": BACKGROUND_URL,
            **session_template_vars(request),
        })

@app.get("/license-status", response_class=HTMLResponse)
async def license_status_page(request: Request):
    session = request.cookies.get("session")
    if not session:
        return RedirectResponse(url="/", status_code=303)
    try:
        data = serializer.loads(session)
        if not session_is_admin(data):
            return RedirectResponse(url="/", status_code=303)
    except Exception:
        return RedirectResponse(url="/", status_code=303)

    return templates.TemplateResponse(
        "license_status.html",
        {
            "request": request,
            "school_name": SCHOOL_NAME,
            "logo_url": LOGO_URL,
            "theme_name": THEME_NAME,
            "background_url": BACKGROUND_URL,
            **get_license_template_context(),
            **session_template_vars(request),
        },
    )

@app.get("/admin/lan-monitor", response_class=HTMLResponse)
async def admin_lan_monitor(request: Request):
    """Admin dashboard: LAN peers, online status, last collect/push times."""
    session = request.cookies.get("session")
    if not session:
        return RedirectResponse(url="/", status_code=303)
    try:
        data = serializer.loads(session)
        if not session_is_admin(data):
            return RedirectResponse(url="/", status_code=303)
    except Exception:
        return RedirectResponse(url="/", status_code=303)

    ensure_primary_node_or_forbidden()
    scan = request.query_params.get("scan") == "1"
    bases = get_monitored_secondary_bases(scan)
    rows = build_secondary_monitor_rows(bases)
    title = "Secondary machines (LAN)"

    return templates.TemplateResponse("lan_monitor.html", {
        "request": request,
        "school_name": SCHOOL_NAME,
        "logo_url": LOGO_URL,
        "theme_name": THEME_NAME,
        "background_url": BACKGROUND_URL,
        "is_primary": True,
        "machine_node_role": NODE_ROLE,
        "machine_id": MACHINE_ID,
        "local_ip": get_local_ip(),
        "scan_lan": scan,
        "peer_rows": rows,
        "peer_title": title,
        "sync_secret_set": bool(SYNC_SECRET),
        "lan_discovery": LAN_DISCOVERY,
        **session_template_vars(request),
    })

@app.get("/internal-shutdown")
async def internal_shutdown(request: Request):
    """Internal endpoint to trigger server shutdown (only from localhost)."""
    if request.client.host != '127.0.0.1':
        raise HTTPException(status_code=403, detail="Access denied")
    global server
    if server:
        server.should_exit = True
    return {"message": "Server signaled for shutdown"}

@app.get("/shutdown")
async def shutdown(request: Request):
    """Endpoint to shutdown the server (admin only)."""
    session = request.cookies.get("session")
    if not session:
        return RedirectResponse(url="/", status_code=303)
    
    try:
        data = serializer.loads(session)
        if not session_is_admin(data):
            raise HTTPException(status_code=403, detail="Admin access required")
    except:
        return RedirectResponse(url="/", status_code=303)
    
    global server
    if server:
        server.should_exit = True
    return {"message": "Server shutting down..."}

def get_local_ip() -> str:
    """Get the local IP address of the machine."""
    try:
        # Create a socket to get the local IP
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
        return local_ip
    except:
        return "127.0.0.1"

def open_browser():
    """Open the browser after a short delay."""
    time.sleep(1.5)  # Wait for server to start
    local_ip = get_local_ip()
    print(f"\nVoting app is running!")
    print(f"Local URL: http://localhost:{APP_PORT}")
    print(f"Network URL: http://{local_ip}:{APP_PORT}")
    print("\nOther devices on the same network can access the app using the Network URL")
    webbrowser.open(f'http://localhost:{APP_PORT}')

def export_results_table_as_image(results, output_path='results_table.png'):
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    table_data = []
    headers = ["Post", "Candidate", "Votes"]
    for post, cand_dict in results.items():
        for cand, count in cand_dict.items():
            table_data.append([post, cand, count])
    num_rows = len(table_data) + 1
    fig_height = min(max(num_rows * 0.5, 3), 20)
    fig, ax = plt.subplots(figsize=(12, fig_height))
    ax.axis('off')
    table = ax.table(
        cellText=table_data,
        colLabels=headers,
        loc='center',
        cellLoc='left',
        colLoc='left',
    )
    table.auto_set_font_size(False)
    table.set_fontsize(12)
    table.auto_set_column_width(col=list(range(len(headers))))
    for (row, col), cell in table.get_celld().items():
        if row == 0:
            cell.set_fontsize(14)
            cell.set_text_props(weight='bold')
            cell.set_facecolor('#cccccc')
    fig.text(
        0.01,
        0.01,
        get_export_branding_line(),
        ha="left",
        va="bottom",
        fontsize=8,
    )
    plt.subplots_adjust(left=0.01, right=0.99, top=0.99, bottom=0.01)
    plt.savefig(output_path, bbox_inches='tight', pad_inches=0.1)
    plt.close()
    print(f'Results exported as table image: {output_path}')

@app.get("/export-results-image")
def export_results_image():
    results = get_results()
    output_path = "school_election_results.png"
    export_results_table_as_image(results, output_path)
    return FileResponse(output_path, media_type="image/png", filename="school_election_results.png")

def export_results_table_as_pdf(results, output_path='school_election_results.pdf'):
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    table_data = []
    headers = ["Post", "Candidate", "Votes"]
    for post, cand_dict in results.items():
        for cand, count in cand_dict.items():
            table_data.append([post, cand, count])
    num_rows = len(table_data) + 1
    fig_height = min(max(num_rows * 0.5, 3), 20)
    fig, ax = plt.subplots(figsize=(12, fig_height))
    ax.axis('off')
    table = ax.table(
        cellText=table_data,
        colLabels=headers,
        loc='center',
        cellLoc='left',
        colLoc='left',
    )
    table.auto_set_font_size(False)
    table.set_fontsize(12)
    table.auto_set_column_width(col=list(range(len(headers))))
    for (row, col), cell in table.get_celld().items():
        if row == 0:
            cell.set_fontsize(14)
            cell.set_text_props(weight='bold')
            cell.set_facecolor('#cccccc')
    fig.text(
        0.01,
        0.01,
        get_export_branding_line(),
        ha="left",
        va="bottom",
        fontsize=8,
    )
    plt.subplots_adjust(left=0.01, right=0.99, top=0.99, bottom=0.01)
    with PdfPages(output_path) as pdf:
        pdf.savefig(fig, bbox_inches='tight', pad_inches=0.1)
    plt.close()
    print(f'Results exported as table PDF: {output_path}')

@app.get("/export-results-pdf")
def export_results_pdf():
    results = get_results()
    output_path = "school_election_results.pdf"
    export_results_table_as_pdf(results, output_path)
    return FileResponse(output_path, media_type="application/pdf", filename="school_election_results.pdf")

def export_results_table_as_reportlab(results, output_path='results_table.pdf'):
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    c = canvas.Canvas(output_path, pagesize=letter)
    width, height = letter
    y = height - 40
    c.setFont("Helvetica-Bold", 16)
    c.drawString(40, y, "Election Results")
    y -= 30
    c.setFont("Helvetica-Bold", 12)
    c.drawString(40, y, "Post")
    c.drawString(200, y, "Candidate")
    c.drawString(400, y, "Votes")
    c.setFont("Helvetica", 12)
    y -= 20
    for post, cand_dict in results.items():
        for cand, count in cand_dict.items():
            c.drawString(40, y, post)
            c.drawString(200, y, cand)
            c.drawString(400, y, str(count))
            y -= 18
            if y < 40:
                c.showPage()
                y = height - 40
    c.setFont("Helvetica", 8)
    c.drawString(40, 20, get_export_branding_line())
    c.save()
    print(f'Results exported as table PDF: {output_path}')

@app.get("/export-results-reportlab")
def export_results_reportlab():
    results = get_results()
    output_path = "results_table.pdf"
    export_results_table_as_reportlab(results, output_path)
    return FileResponse(output_path, media_type="application/pdf", filename="results_table.pdf")

if __name__ == "__main__":
    # Ensure votes file exists before starting the server
    try:
        initialize_votes_file()
        # Ensure candidates file exists
        load_candidates()
        # Create batch files
        create_batch_files()
    except Exception as e:
        print(f"Error initializing files: {e}")
        sys.exit(1)
        
    max_startup_retries = 5
    retry_interval_seconds = 1
    
    lock = filelock.FileLock(LOCK_FILE_PATH)
    
    for attempt in range(max_startup_retries):
        try:
            lock.acquire(timeout=0.1)
            print(f"Acquired lock: {LOCK_FILE_PATH}")
            
            # Import required modules
            import uvicorn
            import logging
            logging.basicConfig(level=logging.INFO)
            
            threading.Thread(target=open_browser, daemon=True).start()
            
            config_uvicorn = uvicorn.Config(app, host=APP_HOST, port=APP_PORT, log_config=None)
            server = uvicorn.Server(config_uvicorn)
            try:
                server.run()
            except KeyboardInterrupt:
                server.should_exit = True
            finally:
                print(f"Releasing lock: {LOCK_FILE_PATH}")
                lock.release()
                if os.path.exists(LOCK_FILE_PATH):
                    try:
                        os.remove(LOCK_FILE_PATH)
                    except:
                        pass
            sys.exit(0)

        except filelock.Timeout:  # Another instance is running
            print(f"Another instance of the app is already running. Attempting to shut it down (attempt {attempt + 1}/{max_startup_retries})...")
            try:
                # Try to connect to the existing instance
                response = requests.get(f'http://localhost:{APP_PORT}/internal-shutdown', timeout=1)
                if response.status_code == 200:
                    print("Existing instance signaled successfully. Waiting for it to release lock...")
                    time.sleep(retry_interval_seconds)
                else:
                    print(f"Failed to signal existing instance (Status: {response.status_code}). Retrying lock acquisition.")
            except requests.exceptions.ConnectionError:
                print("Could not connect to existing instance. It might be stuck or not running properly.")
                # Try to force remove the lock file
                try:
                    if os.path.exists(LOCK_FILE_PATH):
                        os.remove(LOCK_FILE_PATH)
                        print("Force removed lock file. Retrying...")
                except:
                    print("Could not remove lock file. Please close the existing instance manually.")
            except requests.exceptions.Timeout:
                print("Timeout while trying to connect to existing instance. Retrying lock acquisition.")
            except Exception as e:
                print(f"An unexpected error occurred while signaling: {e}. Retrying lock acquisition.")
            
            time.sleep(retry_interval_seconds)
            continue

        except Exception as e:
            print(f"An error occurred during app startup: {e}")
            sys.exit(1)

    # If we get here, we couldn't acquire the lock after retries
    print(f"Failed to acquire application lock after {max_startup_retries} attempts. Another instance might be stuck.")
    print("Please close the existing instance manually and try again.")
    sys.exit(1)

    # Example: export results as table image
    try:
        results = get_results()
        export_results_table_as_image(results, 'school_election_results.png')
    except Exception as e:
        print(f'Could not export results as image: {e}') 